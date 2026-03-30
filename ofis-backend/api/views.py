from django.shortcuts import render
from django.http import JsonResponse
from rest_framework import generics, status, viewsets,permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.parsers import MultiPartParser, FormParser
from django.contrib.auth.models import User
from django.db.models import Sum, Count
from django.utils import timezone
from datetime import timedelta
import os
from django.http import FileResponse
from django.conf import settings
from django.core.files.base import ContentFile
import tempfile
from io import BytesIO
import csv
import qrcode
import base64
from django.db.models.deletion import ProtectedError
from django.db import IntegrityError
from notifications.models import Notification  # si vous utilisez le système de notifications





# Bibliothèques pour génération de fichiers (avec fallback si non installées)
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("⚠️ ReportLab non installé. La génération PDF utilisera le fallback texte.")

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ OpenPyXL non installé. La génération Excel utilisera le fallback CSV.")
from .models import (
    Client, Technician, Manager, Mission, Report, Position, Team,
    GeneratedReport, GeneratedReportDownload,
    Service, Equipment, MissionV2, MissionReport,
    BonDeCommande, LigneBonDeCommande,
    RapportJournalier, RapportHebdomadaire,
    SuiviMedical,
    OrdreTravail,
    RapportProjetCadre,
    RapportHebdoCadre, LigneRapportHebdoCadre
)

from .serializers import (
    ClientSerializer, TechnicianSerializer, ManagerSerializer,
    MissionSerializer, ReportSerializer, UserSerializer,
    PositionSerializer, TeamSerializer,
    GeneratedReportSerializer,
    GeneratedReportCreateSerializer, GeneratedReportUploadSerializer,
    GeneratedReportDownloadSerializer,
    ServiceSerializer, EquipmentSerializer, MissionV2Serializer, MissionReportSerializer,
    BonDeCommandeSerializer, LigneBonDeCommandeSerializer,
    RapportJournalierSerializer, RapportHebdomadaireSerializer,
    SuiviMedicalSerializer,
    OrdreTravailSerializer,
    RapportProjetCadreSerializer,
    RapportHebdoCadreSerializer,
    LigneRapportHebdoCadreSerializer
)

# ===== VIEWSETS EXISTANTS =====

class GeneratedReportViewSet(viewsets.ModelViewSet):
    """ViewSet pour les nouveaux rapports génériques"""
    permission_classes = [IsAuthenticated]
    serializer_class = GeneratedReportSerializer

    def get_queryset(self):
        """Les utilisateurs ne voient que leurs propres rapports"""
        user = self.request.user
        if user.is_superuser:
            return GeneratedReport.objects.all()
        return GeneratedReport.objects.filter(created_by=user)

    def get_serializer_class(self):
        if self.action == 'create':
            return GeneratedReportCreateSerializer
        elif self.action == 'upload':
            return GeneratedReportUploadSerializer
        return GeneratedReportSerializer

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

    @action(detail=True, methods=['post'])
    def generate(self, request, pk=None):
        """Génère un vrai rapport avec fichier"""
        report = self.get_object()
        try:
            report.status = 'en_cours'
            report.save()
            file_content, filename, content_type = self._generate_report_file(report)
            report.file.save(filename, ContentFile(file_content))
            report.status = 'genere'
            report.generated_at = timezone.now()
            report.file_size = report.file.size
            report.original_filename = filename
            report.save()
            return Response({
                'message': 'Rapport généré avec succès',
                'file_url': report.file.url if report.file else None,
                'filename': filename,
                'id': report.id
            })
        except Exception as e:
            report.status = 'erreur'
            report.save()
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _generate_report_file(self, report):
        """Génère le contenu du fichier avec fallback en cas d'erreur"""
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        clean_title = "".join(c for c in report.title if c.isalnum() or c in (' ', '-', '_')).strip()
        clean_title = clean_title.replace(' ', '_')[:30]
        base_filename = f"{clean_title}_{timestamp}"

        title = report.title or "Rapport sans titre"
        description = report.description or ""
        created_at = report.created_at.strftime('%d/%m/%Y %H:%M')
        created_by = report.created_by.username if report.created_by else "Inconnu"
        report_type_display = dict(report.REPORT_TYPES).get(report.report_type, report.report_type)

        try:
            if report.format == 'pdf':
                try:
                    return self._generate_pdf(report, base_filename, title, description, created_at, created_by, report_type_display)
                except Exception as e:
                    print(f"Erreur génération PDF: {e}, fallback vers TXT")
                    return self._generate_text_fallback(report, base_filename, title, description, created_at, created_by, report_type_display, 'pdf')
            elif report.format in ['excel', 'xlsx']:
                try:
                    return self._generate_excel(report, base_filename, title, description, created_at, created_by, report_type_display)
                except Exception as e:
                    print(f"Erreur génération Excel: {e}, fallback vers CSV")
                    return self._generate_csv_fallback(report, base_filename, title, description, created_at, created_by, report_type_display)
            elif report.format == 'csv':
                try:
                    return self._generate_csv(report, base_filename, title, description, created_at, created_by, report_type_display)
                except Exception as e:
                    print(f"Erreur génération CSV: {e}, fallback vers TXT")
                    return self._generate_text_fallback(report, base_filename, title, description, created_at, created_by, report_type_display, 'csv')
            else:
                return self._generate_text(report, base_filename, title, description, created_at, created_by, report_type_display)
        except Exception as e:
            print(f"Erreur générale: {e}, utilisation du fallback ultime")
            return self._generate_text_fallback(report, base_filename, title, description, created_at, created_by, report_type_display, 'txt')

    def _generate_pdf(self, report, base_filename, title, description, created_at, created_by, report_type_display):
        if not REPORTLAB_AVAILABLE:
            raise Exception("ReportLab n'est pas installé")
        filename = base_filename + '.pdf'
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        p.setFont("Helvetica-Bold", 18)
        p.drawString(2*cm, height-3*cm, title[:80])
        p.setFont("Helvetica-Bold", 12)
        p.drawString(2*cm, height-5*cm, "Informations générales:")
        p.setFont("Helvetica", 11)
        p.drawString(3*cm, height-6*cm, f"Créé le: {created_at}")
        p.drawString(3*cm, height-7*cm, f"Créé par: {created_by}")
        p.drawString(3*cm, height-8*cm, f"Type: {report_type_display}")
        p.setFont("Helvetica-Bold", 12)
        p.drawString(2*cm, height-10*cm, "Description:")
        p.setFont("Helvetica", 10)
        y_position = height-11*cm
        words = description.split()
        lines = []
        current_line = ""
        for word in words:
            if len(current_line + " " + word) < 80:
                current_line += (" " + word if current_line else word)
            else:
                lines.append(current_line)
                current_line = word
        if current_line:
            lines.append(current_line)
        for line in lines:
            p.drawString(3*cm, y_position, line)
            y_position -= 0.5*cm
        if report.parameters:
            y_position -= 1*cm
            p.setFont("Helvetica-Bold", 12)
            p.drawString(2*cm, y_position, "Paramètres:")
            y_position -= 0.8*cm
            p.setFont("Helvetica", 10)
            for key, value in report.parameters.items():
                p.drawString(3*cm, y_position, f"{key}: {value}")
                y_position -= 0.5*cm
        p.setFont("Helvetica-Oblique", 8)
        p.drawString(2*cm, 2*cm, f"Généré le {timezone.now().strftime('%d/%m/%Y %H:%M')} - OFIS")
        p.showPage()
        p.save()
        buffer.seek(0)
        return buffer.getvalue(), filename, 'application/pdf'

    def _generate_excel(self, report, base_filename, title, description, created_at, created_by, report_type_display):
        if not OPENPYXL_AVAILABLE:
            raise Exception("OpenPyXL n'est pas installé")
        filename = base_filename + '.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapport"
        ws['A1'] = title
        ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)
        ws.merge_cells('A1:C1')
        ws['A3'] = "Date de création:"
        ws['B3'] = created_at
        ws['A4'] = "Créé par:"
        ws['B4'] = created_by
        ws['A5'] = "Type de rapport:"
        ws['B5'] = report_type_display
        ws['A7'] = "Description:"
        ws['A8'] = description
        ws.merge_cells('A8:C8')
        if report.parameters:
            ws['A10'] = "Paramètres:"
            ws['A10'].font = openpyxl.styles.Font(bold=True)
            row = 11
            for key, value in report.parameters.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = str(value)
                row += 1
        for column in ['A', 'B', 'C']:
            ws.column_dimensions[column].width = 20
        ws['A20'] = f"Généré le {timezone.now().strftime('%d/%m/%Y %H:%M')} - OFIS"
        ws['A20'].font = openpyxl.styles.Font(italic=True, size=8)
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue(), filename, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def _generate_csv(self, report, base_filename, title, description, created_at, created_by, report_type_display):
        filename = base_filename + '.csv'
        buffer = BytesIO()
        writer = csv.writer(buffer)
        writer.writerow([title])
        writer.writerow([])
        writer.writerow(["Date de création:", created_at])
        writer.writerow(["Créé par:", created_by])
        writer.writerow(["Type:", report_type_display])
        writer.writerow([])
        writer.writerow(["Description:"])
        writer.writerow([description])
        if report.parameters:
            writer.writerow([])
            writer.writerow(["Paramètres:"])
            for key, value in report.parameters.items():
                writer.writerow([key, value])
        writer.writerow([])
        writer.writerow([f"Généré le {timezone.now().strftime('%d/%m/%Y %H:%M')} - OFIS"])
        buffer.seek(0)
        return buffer.getvalue(), filename, 'text/csv'

    def _generate_text(self, report, base_filename, title, description, created_at, created_by, report_type_display):
        filename = base_filename + '.txt'
        content = f"""
========================================
{title}
========================================

INFORMATIONS GÉNÉRALES
----------------------------------------
Date de création: {created_at}
Créé par: {created_by}
Type: {report_type_display}

DESCRIPTION
----------------------------------------
{description}

PARAMÈTRES
----------------------------------------
"""
        if report.parameters:
            for key, value in report.parameters.items():
                content += f"{key}: {value}\n"
        else:
            content += "Aucun paramètre\n"
        content += f"\n\nGénéré le {timezone.now().strftime('%d/%m/%Y %H:%M')} - OFIS"
        return content.encode('utf-8'), filename, 'text/plain'

    def _generate_text_fallback(self, report, base_filename, title, description, created_at, created_by, report_type_display, original_format):
        filename = f"{base_filename}_fallback.txt"
        content = f"""
==================================================
⚠️  RAPPORT GÉNÉRÉ EN FORMAT TEXTE (FALLBACK)  ⚠️
==================================================

Le format demandé ({original_format.upper()}) n'a pas pu être généré.
Cause possible: Bibliothèque manquante ou erreur de génération.

------------------------------------------
INFORMATIONS DU RAPPORT
------------------------------------------
Titre: {title}
Date: {created_at}
Créé par: {created_by}
Type: {report_type_display}
Format original demandé: {original_format}

------------------------------------------
DESCRIPTION
------------------------------------------
{description}

------------------------------------------
PARAMÈTRES
------------------------------------------
"""
        if report.parameters:
            for key, value in report.parameters.items():
                content += f"{key}: {value}\n"
        else:
            content += "Aucun paramètre\n"
        content += f"""
------------------------------------------
Généré le {timezone.now().strftime('%d/%m/%Y %H:%M')} - OFIS
==================================================
"""
        return content.encode('utf-8'), filename, 'text/plain'

    def _generate_csv_fallback(self, report, base_filename, title, description, created_at, created_by, report_type_display):
        filename = base_filename + '.csv'
        buffer = BytesIO()
        writer = csv.writer(buffer)
        writer.writerow(["RAPPORT", title])
        writer.writerow([])
        writer.writerow(["Date création", created_at])
        writer.writerow(["Créé par", created_by])
        writer.writerow(["Type", report_type_display])
        writer.writerow([])
        writer.writerow(["Description"])
        writer.writerow([description])
        if report.parameters:
            writer.writerow([])
            writer.writerow(["Paramètres"])
            for key, value in report.parameters.items():
                writer.writerow([key, value])
        buffer.seek(0)
        return buffer.getvalue(), filename, 'text/csv'

    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        report = self.get_object()
        if not report.file:
            return Response({'error': 'Aucun fichier disponible'}, status=status.HTTP_404_NOT_FOUND)
        try:
            if not os.path.exists(report.file.path):
                return Response({'error': 'Fichier non trouvé'}, status=status.HTTP_404_NOT_FOUND)
            report.download_count += 1
            report.save()
            GeneratedReportDownload.objects.create(
                report=report,
                downloaded_by=request.user,
                ip_address=self._get_client_ip(request)
            )
            filename = report.original_filename or report.filename or f"rapport_{report.id}.pdf"
            response = FileResponse(report.file.open('rb'), as_attachment=True, filename=filename)
            return response
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='upload')
    def upload(self, request):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            report = serializer.save()
            return Response(GeneratedReportSerializer(report, context=self.get_serializer_context()).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        queryset = self.get_queryset()
        now = timezone.now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        week_start = today_start - timedelta(days=today_start.weekday())
        month_start = today_start.replace(day=1)
        stats = {
            'total_reports': queryset.count(),
            'reports_this_month': queryset.filter(created_at__gte=month_start).count(),
            'reports_this_week': queryset.filter(created_at__gte=week_start).count(),
            'generated_today': queryset.filter(generated_at__gte=today_start, status='genere').count(),
            'downloads_total': GeneratedReportDownload.objects.filter(report__in=queryset).count(),
            'by_type': dict(queryset.values('report_type').annotate(count=Count('id')).values_list('report_type', 'count')),
        }
        return Response(stats)

    def _get_client_ip(self, request):
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            return x_forwarded_for.split(',')[0]
        return request.META.get('REMOTE_ADDR')


# ===== REPORT VIEWS EXISTANTS =====
class ReportViewSet(viewsets.ModelViewSet):
    queryset = Report.objects.all().order_by('-reportDate')
    serializer_class = ReportSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save()


class ReportUploadView(generics.CreateAPIView):
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)
    serializer_class = ReportSerializer

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'Aucun fichier fourni'}, status=400)
        if file.size > 10 * 1024 * 1024:
            return Response({'error': 'Fichier trop grand (max 10MB)'}, status=400)
        report = Report.objects.create(
            title=request.data.get('title', file.name),
            description=request.data.get('description', ''),
            file=file,
            created_by=request.user
        )
        serializer = ReportSerializer(report)
        return Response(serializer.data, status=201)


class ReportDeleteView(generics.DestroyAPIView):
    queryset = Report.objects.all()
    permission_classes = [IsAuthenticated]

    def perform_destroy(self, instance):
        if instance.file and os.path.exists(instance.file.path):
            os.remove(instance.file.path)
        instance.delete()


# ===== POSITION VIEWS =====
class PositionListCreateView(generics.ListCreateAPIView):
    serializer_class = PositionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        team_id = self.request.query_params.get('team')
        if team_id:
            return Position.objects.filter(team_id=team_id)
        return Position.objects.all()[:100]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class TeamLatestPositionView(generics.RetrieveAPIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, team_id):
        try:
            latest = Position.objects.filter(team_id=team_id).first()
            if latest:
                serializer = PositionSerializer(latest)
                return Response(serializer.data)
            return Response({'detail': 'Aucune position trouvée'}, status=404)
        except Exception as e:
            return Response({'error': str(e)}, status=400)


# ===== HELLO WORLD =====
def hello_world(request):
    return JsonResponse({"message": "Hello from OFIS API!"})


# ===== CLIENTS =====
class ClientListCreate(generics.ListCreateAPIView):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer
    permission_classes = [IsAuthenticated]

class ClientRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer
    permission_classes = [IsAuthenticated]

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            instance.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except (ProtectedError, IntegrityError):
            return Response(
                {"detail": "Ce client ne peut pas être supprimé car il est référencé par d'autres enregistrements (missions, bons de commande, ordres de travail, rapports, etc.)."},
                status=status.HTTP_400_BAD_REQUEST
            )


# ===== TECHNICIENS =====
class TechnicianListCreate(generics.ListCreateAPIView):
    queryset = Technician.objects.all()
    serializer_class = TechnicianSerializer
    permission_classes = [IsAuthenticated]


# ===== MANAGERS =====
class ManagerListCreate(generics.ListCreateAPIView):
    queryset = Manager.objects.all()
    serializer_class = ManagerSerializer
    permission_classes = [IsAuthenticated]


# ===== MISSIONS =====
class MissionListCreate(generics.ListCreateAPIView):
    queryset = Mission.objects.all()
    serializer_class = MissionSerializer
    permission_classes = [IsAuthenticated]


# ===== UTILISATEURS =====
class UserListCreate(generics.ListCreateAPIView):
    queryset = User.objects.all().order_by('username')
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]


class UserRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]


# ===== TECHNICIEN DETAIL VIEWS =====
class TechnicianRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Technician.objects.all()
    serializer_class = TechnicianSerializer
    permission_classes = [IsAuthenticated]


# ===== MANAGER DETAIL VIEWS =====
class ManagerRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Manager.objects.all()
    serializer_class = ManagerSerializer
    permission_classes = [IsAuthenticated]


# ===== MISSION DETAIL VIEWS =====
class MissionRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Mission.objects.all()
    serializer_class = MissionSerializer
    permission_classes = [IsAuthenticated]


# ===== TEAM VIEWS =====
class TeamListCreate(generics.ListCreateAPIView):
    queryset = Team.objects.all()
    serializer_class = TeamSerializer
    permission_classes = [IsAuthenticated]


class TeamRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Team.objects.all()
    serializer_class = TeamSerializer
    permission_classes = [IsAuthenticated]


# ========== NOUVELLES VUES ADMIN DATABASE CORRIGÉES ==========
import os
from django.db import connection
from django.http import HttpResponse
from datetime import datetime
import glob
import json

class AdminTablesView(generics.GenericAPIView):
    permission_classes = [IsAdminUser]
    def get(self, request):
        tables = []
        with connection.cursor() as cursor:
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';")
            for table in cursor.fetchall():
                table_name = table[0]
                cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
                count = cursor.fetchone()[0]
                tables.append({'name': table_name, 'count': count})
        return Response(tables)


class AdminBackupView(generics.GenericAPIView):
    permission_classes = [IsAdminUser]
    def post(self, request):
        try:
            backup_dir = os.path.join(settings.BASE_DIR, 'backups')
            os.makedirs(backup_dir, exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_file = os.path.join(backup_dir, f'backup_{timestamp}.sql')
            with open(backup_file, 'w', encoding='utf-8') as f:
                for line in connection.connection.iterdump():
                    f.write('%s\n' % line)
            response = FileResponse(open(backup_file, 'rb'), as_attachment=True, filename=f'backup_{timestamp}.sql')
            return response
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class AdminRestoreView(generics.GenericAPIView):
    permission_classes = [IsAdminUser]
    parser_classes = [MultiPartParser, FormParser]
    def post(self, request):
        try:
            backup_file = request.FILES.get('backup')
            if not backup_file:
                return Response({'error': 'Aucun fichier fourni'}, status=400)
            content = backup_file.read().decode('utf-8')
            connection.close()
            with connection.cursor() as cursor:
                cursor.executescript(content)
            return Response({'message': 'Base de données restaurée avec succès'})
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class AdminOptimizeView(generics.GenericAPIView):
    permission_classes = [IsAdminUser]
    def post(self, request):
        try:
            with connection.cursor() as cursor:
                cursor.execute("VACUUM;")
            return Response({'message': 'Base de données optimisée avec succès'})
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class AdminBackupHistoryView(generics.GenericAPIView):
    permission_classes = [IsAdminUser]
    def get(self, request):
        backup_dir = os.path.join(settings.BASE_DIR, 'backups')
        os.makedirs(backup_dir, exist_ok=True)
        backups = []
        for file in glob.glob(os.path.join(backup_dir, '*.sql')):
            stat = os.stat(file)
            backups.append({
                'filename': os.path.basename(file),
                'date': datetime.fromtimestamp(stat.st_ctime).strftime('%d/%m/%Y %H:%M'),
                'size': f"{stat.st_size / 1024:.2f} KB"
            })
        backups.sort(key=lambda x: x['date'], reverse=True)
        return Response(backups)


class AdminTableView(generics.GenericAPIView):
    permission_classes = [IsAdminUser]
    def get(self, request, table_name):
        try:
            with connection.cursor() as cursor:
                cursor.execute(f"SELECT * FROM {table_name} LIMIT 100")
                columns = [col[0] for col in cursor.description]
                rows = cursor.fetchall()
                data = []
                for row in rows:
                    item = {}
                    for i, col in enumerate(columns):
                        item[col] = row[i]
                    data.append(item)
                return Response(data)
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class AdminTableExportView(generics.GenericAPIView):
    permission_classes = [IsAdminUser]
    def get(self, request, table_name):
        try:
            import csv
            from io import StringIO
            with connection.cursor() as cursor:
                cursor.execute(f"SELECT * FROM {table_name}")
                columns = [col[0] for col in cursor.description]
                rows = cursor.fetchall()
                output = StringIO()
                writer = csv.writer(output)
                writer.writerow(columns)
                writer.writerows(rows)
                response = HttpResponse(output.getvalue(), content_type='text/csv')
                response['Content-Disposition'] = f'attachment; filename="{table_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
                return response
        except Exception as e:
            return Response({'error': str(e)}, status=500)


# ===== NOUVELLES VUES POUR LA GESTION DES ÉQUIPES =====

class ServiceViewSet(viewsets.ModelViewSet):
    queryset = Service.objects.all()
    serializer_class = ServiceSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=True, methods=['get'])
    def teams(self, request, pk=None):
        service = self.get_object()
        teams = service.teams.all()
        serializer = TeamSerializer(teams, many=True, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def equipment(self, request, pk=None):
        service = self.get_object()
        equipment = service.equipment.all()
        serializer = EquipmentSerializer(equipment, many=True, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def missions(self, request, pk=None):
        service = self.get_object()
        missions = MissionV2.objects.filter(service=service)
        serializer = MissionV2Serializer(missions, many=True, context={'request': request})
        return Response(serializer.data)


class EquipmentViewSet(viewsets.ModelViewSet):
    queryset = Equipment.objects.all()
    serializer_class = EquipmentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = Equipment.objects.all()
        service = self.request.query_params.get('service')
        team = self.request.query_params.get('team')
        if service:
            queryset = queryset.filter(service__name=service)
        if team:
            queryset = queryset.filter(team_id=team)
        return queryset


# ===== MISSION V2 VIEWSET CORRIGÉ =====
class MissionV2ViewSet(viewsets.ModelViewSet):
    queryset = MissionV2.objects.all()
    serializer_class = MissionV2Serializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = MissionV2.objects.all()
        status = self.request.query_params.get('status')
        service = self.request.query_params.get('service')
        team = self.request.query_params.get('team')
        if status:
            queryset = queryset.filter(status=status)
        if service:
            queryset = queryset.filter(service__name=service)
        if team:
            queryset = queryset.filter(team_id=team)
        return queryset

    def perform_create(self, serializer):
        mission = serializer.save(created_by=self.request.user)
        print(f"✅ Mission créée: {mission.title}")
        print(f"👥 Techniciens assignés: {mission.assigned_to.all()}")

    @action(detail=True, methods=['post'])
    def start(self, request, pk=None):
        mission = self.get_object()
        mission.status = 'en_cours'
        mission.save()
        return Response({'status': 'Mission démarrée', 'mission_id': mission.id})

    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        mission = self.get_object()
        mission.status = 'terminee'
        mission.actual_hours = request.data.get('actual_hours', mission.estimated_hours)
        mission.save()
        return Response({'status': 'Mission terminée', 'mission_id': mission.id})

    @action(detail=True, methods=['post'])
    def assign(self, request, pk=None):
        mission = self.get_object()
        user_ids = request.data.get('user_ids', [])
        mission.assigned_to.clear()
        for user_id in user_ids:
            try:
                user = User.objects.get(id=user_id)
                mission.assigned_to.add(user)
                print(f"✅ Technicien {user.username} assigné")
            except User.DoesNotExist:
                print(f"❌ Utilisateur {user_id} non trouvé")
        return Response({
            'status': 'Techniciens assignés',
            'assigned_to_ids': [user.id for user in mission.assigned_to.all()]
        })


class MissionReportViewSet(viewsets.ModelViewSet):
    queryset = MissionReport.objects.all()
    serializer_class = MissionReportSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = MissionReport.objects.all()
        mission_id = self.request.query_params.get('mission')
        if mission_id:
            queryset = queryset.filter(mission_id=mission_id)
        return queryset

    def perform_create(self, serializer):
        serializer.save(technician=self.request.user)


# ===== DASHBOARD STATS VIEW (MODIFIÉ) =====
from .models import OrdreTravail, SuiviOT, RapportJournalier, RapportHebdoCadre, Client
from django.db.models import Sum
from datetime import date, timedelta
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

class DashboardStatsView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        is_staff = user.is_staff or user.is_superuser

        # OT
        ots = OrdreTravail.objects.all()
        if not is_staff:
            ots = ots.filter(technicien=user)
        ot_counts = {
            'planifie': ots.filter(statut='planifie').count(),
            'en_cours': ots.filter(statut='en_cours').count(),
            'a_valider': ots.filter(statut='termine', statut_validation='en_attente').count(),
            'valide': ots.filter(statut_validation='valide').count(),
            'rejete': ots.filter(statut_validation='rejete').count(),
        }

        # Heures
        suivis = SuiviOT.objects.all()
        if not is_staff:
            suivis = suivis.filter(technicien=user)
        total_heures = suivis.aggregate(total=Sum('heures'))['total'] or 0

        # Rapports journaliers
        rapports_journaliers = RapportJournalier.objects.all()
        if not is_staff:
            rapports_journaliers = rapports_journaliers.filter(technicien=user)
        total_rapports_journaliers = rapports_journaliers.count()

        # Rapports hebdomadaires
        rapports_hebdo = RapportHebdoCadre.objects.all()
        if not is_staff:
            rapports_hebdo = rapports_hebdo.filter(cadre=user)
        total_rapports_hebdo = rapports_hebdo.count()

        # Clients (seulement pour staff)
        total_clients = Client.objects.count() if is_staff else 0

        # Évolution des heures sur 7 jours
        today = date.today()
        last_7_days = [today - timedelta(days=i) for i in range(6, -1, -1)]
        heures_par_jour = []
        for d in last_7_days:
            total = SuiviOT.objects.filter(date=d).aggregate(total=Sum('heures'))['total'] or 0
            if not is_staff:
                total = SuiviOT.objects.filter(date=d, technicien=user).aggregate(total=Sum('heures'))['total'] or 0
            heures_par_jour.append({'date': d.strftime('%Y-%m-%d'), 'heures': float(total)})

        return Response({
            'ot_counts': ot_counts,
            'total_heures': total_heures,
            'total_rapports_journaliers': total_rapports_journaliers,
            'total_rapports_hebdo': total_rapports_hebdo,
            'total_clients': total_clients,
            'heures_par_jour': heures_par_jour,
        })

# ===== BONS DE COMMANDE =====
class BonDeCommandeViewSet(viewsets.ModelViewSet):
    queryset = BonDeCommande.objects.all().order_by('-date_creation')
    serializer_class = BonDeCommandeSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        statut = self.request.query_params.get('statut')
        if statut:
            queryset = queryset.filter(statut=statut)
        return queryset

    @action(detail=True, methods=['get'])
    def qr(self, request, pk=None):
        """Génère et retourne l'image QR du bon"""
        bon = self.get_object()
        validation_url = f"http://localhost:3000/valider-bon/{bon.qr_code}"
        qr = qrcode.make(validation_url)
        buffer = BytesIO()
        qr.save(buffer, format='PNG')
        img_str = base64.b64encode(buffer.getvalue()).decode()
        return Response({'qr_code': f"data:image/png;base64,{img_str}"})

    @action(detail=False, methods=['post'], url_path='valider/(?P<qr_code>[^/.]+)')
    def valider_par_qr(self, request, qr_code=None):
        """Valide un bon à partir de son qr_code (scanné)"""
        try:
            bon = BonDeCommande.objects.get(qr_code=qr_code)
            bon.statut = 'valide'
            bon.date_validation = timezone.now()
            bon.save()
            return Response({'status': 'Bon validé', 'bon': self.get_serializer(bon).data})
        except BonDeCommande.DoesNotExist:
            return Response({'error': 'Bon non trouvé'}, status=404)

    @action(detail=True, methods=['post'])
    def valider(self, request, pk=None):
        """Valider un bon (changer statut à 'valide')"""
        bon = self.get_object()
        bon.statut = 'valide'
        bon.date_validation = timezone.now()
        bon.save()
        return Response({'status': 'Bon validé', 'bon': self.get_serializer(bon).data})

    @action(detail=True, methods=['post'])
    def rejeter(self, request, pk=None):
        """Rejeter un bon (changer statut à 'annule')"""
        bon = self.get_object()
        bon.statut = 'annule'
        bon.save()
        return Response({'status': 'Bon rejeté', 'bon': self.get_serializer(bon).data})


# ===== RAPPORTS JOURNALIERS, HEBDOMADAIRES, PROJET =====

class RapportJournalierViewSet(viewsets.ModelViewSet):
    queryset = RapportJournalier.objects.all()
    serializer_class = RapportJournalierSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.is_staff or user.is_superuser:
            return RapportJournalier.objects.all()
        return RapportJournalier.objects.filter(technicien=user)

    def perform_create(self, serializer):
        serializer.save(technicien=self.request.user)

class RapportHebdoCadreViewSet(viewsets.ModelViewSet):
    serializer_class = RapportHebdoCadreSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user

        # Admin / staff voit tout
        if user.is_staff or user.is_superuser:
            return RapportHebdoCadre.objects.all().order_by('-created_at')

        # Sinon l'utilisateur ne voit que ses rapports
        return RapportHebdoCadre.objects.filter(cadre=user).order_by('-created_at')

    def perform_create(self, serializer):
        serializer.save(cadre=self.request.user)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(cadre=request.user)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def ajouter_ligne(self, request, pk=None):
        rapport = self.get_object()
        serializer = LigneRapportHebdoCadreSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save(rapport=rapport)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['put'], url_path='lignes/(?P<ligne_id>[^/.]+)')
    def modifier_ligne(self, request, pk=None, ligne_id=None):
        rapport = self.get_object()

        try:
            ligne = rapport.lignes.get(id=ligne_id)
        except LigneRapportHebdoCadre.DoesNotExist:
            return Response({"detail": "Ligne introuvable"}, status=status.HTTP_404_NOT_FOUND)

        serializer = LigneRapportHebdoCadreSerializer(ligne, data=request.data, partial=True)

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['delete'], url_path='lignes/(?P<ligne_id>[^/.]+)')
    def supprimer_ligne(self, request, pk=None, ligne_id=None):
        rapport = self.get_object()

        try:
            ligne = rapport.lignes.get(id=ligne_id)
        except LigneRapportHebdoCadre.DoesNotExist:
            return Response({"detail": "Ligne introuvable"}, status=status.HTTP_404_NOT_FOUND)

        ligne.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

# ===== RAPPORTS PROJET (pour les cadres) =====



# ===== SUIVI MÉDICAL =====
class SuiviMedicalViewSet(viewsets.ModelViewSet):
    queryset = SuiviMedical.objects.all().order_by('nom')
    serializer_class = SuiviMedicalSerializer
    permission_classes = [IsAuthenticated]


# ===== ORDRES DE TRAVAIL =====
from rest_framework.parsers import MultiPartParser, FormParser
from .models import OrdreTravail, DocumentOT
from .serializers import OrdreTravailSerializer, DocumentOTSerializer

class OrdreTravailViewSet(viewsets.ModelViewSet):
    queryset = OrdreTravail.objects.all().order_by('-date_creation')
    serializer_class = OrdreTravailSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.is_staff or user.is_superuser:
            return OrdreTravail.objects.all()
        return OrdreTravail.objects.filter(technicien=user)

    @action(detail=True, methods=['post'])
    def demarrer(self, request, pk=None):
        ot = self.get_object()
        if ot.statut != 'planifie':
            return Response({'error': 'Cet OT ne peut pas être démarré'}, status=400)
        ot.statut = 'en_cours'
        ot.date_debut = timezone.now()
        ot.save()
        return Response({'status': 'OT démarré', 'ot': self.get_serializer(ot).data})

    @action(detail=True, methods=['post'])
    def terminer(self, request, pk=None):
        ot = self.get_object()
        if ot.statut != 'en_cours':
            return Response({'error': 'Seul un OT en cours peut être terminé'}, status=400)
        serializer = self.get_serializer(ot, data=request.data, partial=True)
        if serializer.is_valid():
            ot = serializer.save()
            ot.statut = 'termine'
            ot.date_fin = timezone.now()
            ot.statut_validation = 'en_attente'
            ot.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)

    @action(detail=True, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def upload_document(self, request, pk=None):
        ot = self.get_object()
        # Seule l'assistante (staff) peut uploader des documents après clôture technique
        if not request.user.is_staff:
            return Response({'error': 'Non autorisé'}, status=403)
        type_doc = request.data.get('type')
        fichier = request.FILES.get('fichier')
        if not type_doc or not fichier:
            return Response({'error': 'Type et fichier requis'}, status=400)
        document = DocumentOT.objects.create(
            ot=ot,
            type=type_doc,
            fichier=fichier,
            nom=fichier.name
        )
        serializer = DocumentOTSerializer(document)
        return Response(serializer.data, status=201)

    @action(detail=True, methods=['post'])
    def valider(self, request, pk=None):
        ot = self.get_object()
        if not request.user.is_staff:
            return Response({'error': 'Non autorisé'}, status=403)
        if ot.statut != 'termine':
            return Response({'error': 'Seul un OT terminé peut être validé'}, status=400)
        ot.statut_validation = 'valide'
        ot.valide_par = request.user
        ot.date_validation = timezone.now()
        ot.date_archivage = timezone.now()  # Archivage automatique
        ot.save()
        return Response({'status': 'OT validé et archivé'})

    @action(detail=True, methods=['post'])
    def rejeter(self, request, pk=None):
        ot = self.get_object()
        if not request.user.is_staff:
            return Response({'error': 'Non autorisé'}, status=403)
        if ot.statut != 'termine':
            return Response({'error': 'Seul un OT terminé peut être rejeté'}, status=400)
        ot.statut_validation = 'rejete'
        ot.save()
        return Response({'status': 'OT rejeté'})


from django.contrib.auth.models import Group

class RapportProjetCadreViewSet(viewsets.ModelViewSet):
    serializer_class = RapportProjetCadreSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        # Les membres du staff (admin, manager) voient tous les rapports
        if user.is_staff or user.is_superuser:
            return RapportProjetCadre.objects.all().order_by('-date_creation')
        # Sinon, le cadre ne voit que ses propres rapports
        return RapportProjetCadre.objects.filter(cadre=user).order_by('-date_creation')

    def perform_create(self, serializer):
        user = self.request.user
        # Vérifier si l'utilisateur est dans le groupe "Cadres" ou est staff (pour debug)
        if not user.is_staff and not user.groups.filter(name='Cadres').exists():
            raise PermissionDenied("Seuls les cadres peuvent créer des rapports projet.")
        serializer.save(cadre=user)





class RapportHebdoCadreViewSet(viewsets.ModelViewSet):
    serializer_class = RapportHebdoCadreSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.is_staff or user.is_superuser:
            return RapportHebdoCadre.objects.all().order_by('-date_debut')
        return RapportHebdoCadre.objects.filter(cadre=user).order_by('-date_debut')

    def perform_create(self, serializer):
        serializer.save(cadre=self.request.user)

    @action(detail=True, methods=['post'])
    def ajouter_ligne(self, request, pk=None):
        rapport = self.get_object()
        serializer = LigneRapportHebdoCadreSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(rapport=rapport)
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)

    @action(detail=True, methods=['put'], url_path='lignes/(?P<ligne_id>[^/.]+)')
    def modifier_ligne(self, request, pk=None, ligne_id=None):
        rapport = self.get_object()
        try:
            ligne = rapport.lignes.get(id=ligne_id)
        except LigneRapportHebdoCadre.DoesNotExist:
            return Response({'error': 'Ligne non trouvée'}, status=404)
        serializer = LigneRapportHebdoCadreSerializer(ligne, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)

    @action(detail=True, methods=['delete'], url_path='lignes/(?P<ligne_id>[^/.]+)')
    def supprimer_ligne(self, request, pk=None, ligne_id=None):
        rapport = self.get_object()
        try:
            ligne = rapport.lignes.get(id=ligne_id)
        except LigneRapportHebdoCadre.DoesNotExist:
            return Response({'error': 'Ligne non trouvée'}, status=404)
        ligne.delete()
        return Response({'status': 'Ligne supprimée'}, status=204)






from .models import SuiviOT
from .serializers import SuiviOTSerializer

class SuiviOTViewSet(viewsets.ModelViewSet):
    serializer_class = SuiviOTSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = SuiviOT.objects.all()
        if not (user.is_staff or user.is_superuser):
            queryset = queryset.filter(technicien=user)
        ot_id = self.request.query_params.get('ot')
        if ot_id:
            queryset = queryset.filter(ot_id=ot_id)
        return queryset

    def perform_create(self, serializer):
        serializer.save(technicien=self.request.user)